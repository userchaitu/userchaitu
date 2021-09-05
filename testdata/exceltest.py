import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\user\\Desktop\\Selenium\\Data\\python_excel_DD.xlsx")
Dict = {}
sheet = workbook.active
print(sheet.title)
cell = sheet.cell(row=1, column=1).value
print(cell)
# value = sheet.cell(row=2, column=2).value = "Chaitu"
# print(value)
rows = sheet.max_row
columns = sheet.max_column
print(rows)
print(columns)

for i in range(1, rows + 1):
    if sheet.cell(row=i, column=1).value == "TC_02":
        for j in range(2, columns + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=3, column=j).value
print(Dict)
