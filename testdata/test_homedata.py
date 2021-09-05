import openpyxl


class HomeTestData:

    test_home_data = [{'First Name': 'Chaitanya Velpuri', 'EmailID': 'abc@abc.com', 'Password': 'mypassword'},
                      {'First Name': 'Ravalika Beeram', 'EmailID': 'rbr@abc.com', 'Password': 'mypassword1'}]

    @staticmethod
    def testdata(test_case):
        workbook = openpyxl.load_workbook("C:\\Users\\user\\Desktop\\Selenium\\Data\\python_excel_DD.xlsx")
        sheet = workbook.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=3, column=j).value
        return [Dict]

